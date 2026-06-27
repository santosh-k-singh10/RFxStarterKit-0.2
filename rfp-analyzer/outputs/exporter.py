"""
outputs/exporter.py
-------------------
Exports the final requirements list to:
  1. Excel (.xlsx)  — colour-coded by category, auto-sized columns,
                      summary sheet + traceability matrix sheet
  2. Markdown (.md) — structured report for sharing / docs

Usage
-----
    from outputs.exporter import export_excel, export_markdown

    export_excel(requirements, "rfp_analysis.xlsx")
    export_markdown(requirements, "rfp_analysis.md", rfp_title="Project Alpha")
"""

from __future__ import annotations

import json
import os
from datetime import datetime
from pathlib import Path
from typing import Optional

import structlog
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.schemas import Category, Priority, Requirement

log = structlog.get_logger()

# ─────────────────────────────────────────────────────────────────────────────
# Colour palette (ARGB format for openpyxl)
# ─────────────────────────────────────────────────────────────────────────────

CATEGORY_FILL = {
    Category.FUNCTIONAL:     PatternFill("solid", fgColor="D6EAF8"),   # Light blue
    Category.NON_FUNCTIONAL: PatternFill("solid", fgColor="D5F5E3"),   # Light green
    Category.COMPLIANCE:     PatternFill("solid", fgColor="FDEBD0"),   # Light orange
    Category.AMBIGUITY:      PatternFill("solid", fgColor="FDEDEC"),   # Light red
    Category.RISK:           PatternFill("solid", fgColor="F9EBEA"),   # Light pink
}

PRIORITY_FILL = {
    Priority.MUST:   PatternFill("solid", fgColor="FADBD8"),   # Red tint
    Priority.SHOULD: PatternFill("solid", fgColor="FDEBD0"),   # Orange tint
    Priority.COULD:  PatternFill("solid", fgColor="EBF5FB"),   # Blue tint
    Priority.WONT:   PatternFill("solid", fgColor="F2F3F4"),   # Grey
}

HEADER_FILL = PatternFill("solid", fgColor="2C3E50")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BOLD_FONT   = Font(bold=True, size=11)
WRAP_ALIGN  = Alignment(wrap_text=True, vertical="top")


# ─────────────────────────────────────────────────────────────────────────────
# Excel export
# ─────────────────────────────────────────────────────────────────────────────

REQUIREMENTS_COLUMNS = [
    ("ID",                    12),
    ("Category",              15),
    ("Title",                 35),
    ("Description",           60),
    ("Source Section",        25),
    ("Page",                   8),
    ("Priority",              10),
    ("Confidence",            12),
    ("Ambiguous?",            12),
    ("Clarification Question",50),
    ("Related IDs",           20),
    ("SAP Modules",           25),
]


def export_excel(requirements: list[Requirement], out_path: str) -> str:
    """
    Write requirements to a colour-coded Excel workbook.

    Sheets created:
      1. Requirements  — main data table
      2. Summary       — count by category and priority
      3. Clarifications — all ambiguous items + their questions

    Returns the absolute path to the written file.
    """
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    _write_requirements_sheet(wb, requirements)
    _write_summary_sheet(wb, requirements)
    _write_clarifications_sheet(wb, requirements)

    # Remove default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(out_path)
    log.info("excel_exported", path=out_path, count=len(requirements))
    return str(Path(out_path).resolve())


def _write_requirements_sheet(wb: Workbook, requirements: list[Requirement]) -> None:
    ws = wb.create_sheet("Requirements", 0)

    # Header row
    headers = [col[0] for col in REQUIREMENTS_COLUMNS]
    widths  = [col[1] for col in REQUIREMENTS_COLUMNS]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_ALIGN

    # Freeze header row
    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, req in enumerate(requirements, start=2):
        fill = CATEGORY_FILL.get(req.category)

        # Get SAP modules if available
        sap_modules_str = ""
        if hasattr(req, 'sap_modules') and req.sap_modules:
            sap_modules_str = ", ".join(req.sap_modules)
        
        values = [
            req.id,
            req.category.value.replace("_", " ").title(),
            req.title,
            req.description,
            req.source_section,
            req.page_ref or "",
            req.priority.value.upper(),
            f"{req.confidence:.0%}",
            "Yes" if req.ambiguity_flag else "No",
            req.clarification_question or "",
            ", ".join(req.related_ids),
            sap_modules_str,
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = WRAP_ALIGN
            if fill:
                cell.fill = fill

    # Column widths
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Row heights (allow wrapped text to breathe)
    for row_idx in range(2, len(requirements) + 2):
        ws.row_dimensions[row_idx].height = 45


def _write_summary_sheet(wb: Workbook, requirements: list[Requirement]) -> None:
    ws = wb.create_sheet("Summary")

    ws["A1"] = "RFP Analysis Summary"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, size=10)

    # Category breakdown
    ws["A4"] = "By Category"
    ws["A4"].font = BOLD_FONT
    ws["B4"] = "Total"
    ws["B4"].font = BOLD_FONT
    ws["C4"] = "Must"
    ws["C4"].font = BOLD_FONT
    ws["D4"] = "Ambiguous"
    ws["D4"].font = BOLD_FONT

    row = 5
    for cat in Category:
        cat_reqs = [r for r in requirements if r.category == cat]
        musts    = sum(1 for r in cat_reqs if r.priority == Priority.MUST)
        ambig    = sum(1 for r in cat_reqs if r.ambiguity_flag)
        ws[f"A{row}"] = cat.value.replace("_", " ").title()
        ws[f"B{row}"] = len(cat_reqs)
        ws[f"C{row}"] = musts
        ws[f"D{row}"] = ambig
        fill = CATEGORY_FILL.get(cat)
        if fill:
            for col in ("A", "B", "C", "D"):
                ws[f"{col}{row}"].fill = fill
        row += 1

    ws[f"A{row}"] = "TOTAL"
    ws[f"A{row}"].font = BOLD_FONT
    ws[f"B{row}"] = len(requirements)
    ws[f"B{row}"].font = BOLD_FONT

    for col in ("A", "B", "C", "D"):
        ws.column_dimensions[col].width = 20


def _write_clarifications_sheet(wb: Workbook, requirements: list[Requirement]) -> None:
    ws = wb.create_sheet("Clarifications Needed")

    ambiguous = [r for r in requirements if r.ambiguity_flag and r.clarification_question]
    if not ambiguous:
        ws["A1"] = "No clarifications needed."
        return

    headers = ["ID", "Section", "Issue", "Question to Ask"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    ws.freeze_panes = "A2"

    for row_idx, req in enumerate(ambiguous, start=2):
        ws.cell(row=row_idx, column=1, value=req.id).alignment = WRAP_ALIGN
        ws.cell(row=row_idx, column=2, value=req.source_section).alignment = WRAP_ALIGN
        ws.cell(row=row_idx, column=3, value=req.title).alignment = WRAP_ALIGN
        ws.cell(row=row_idx, column=4, value=req.clarification_question).alignment = WRAP_ALIGN
        for col in range(1, 5):
            ws.cell(row=row_idx, column=col).fill = CATEGORY_FILL.get(req.category, CATEGORY_FILL[Category.AMBIGUITY])
        ws.row_dimensions[row_idx].height = 50

    for col, width in zip(("A", "B", "C", "D"), (10, 25, 35, 70)):
        ws.column_dimensions[col].width = width


# ─────────────────────────────────────────────────────────────────────────────
# Markdown export
# ─────────────────────────────────────────────────────────────────────────────

def export_markdown(
    requirements: list[Requirement],
    out_path: str,
    rfp_title: str = "RFP Analysis",
) -> str:
    """Write a structured Markdown report."""
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)

    lines: list[str] = []
    lines.append(f"# {rfp_title}")
    lines.append(f"\n_Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}_\n")

    # Summary stats
    lines.append("## Summary\n")
    lines.append("| Category | Count | Must | Ambiguous |")
    lines.append("|---|---|---|---|")
    for cat in Category:
        cat_reqs = [r for r in requirements if r.category == cat]
        musts    = sum(1 for r in cat_reqs if r.priority == Priority.MUST)
        ambig    = sum(1 for r in cat_reqs if r.ambiguity_flag)
        lines.append(f"| {cat.value.replace('_', ' ').title()} | {len(cat_reqs)} | {musts} | {ambig} |")
    lines.append(f"| **Total** | **{len(requirements)}** | | |\n")

    # Per-category sections
    for cat in Category:
        cat_reqs = sorted(
            [r for r in requirements if r.category == cat],
            key=lambda r: r.id,
        )
        if not cat_reqs:
            continue

        lines.append(f"\n## {cat.value.replace('_', ' ').title()}\n")

        for req in cat_reqs:
            conf_pct = f"{req.confidence:.0%}"
            ambig_str = " ⚠ Ambiguous" if req.ambiguity_flag else ""
            lines.append(f"### {req.id} — {req.title}{ambig_str}")
            lines.append(f"- **Priority:** {req.priority.value.upper()}")
            lines.append(f"- **Confidence:** {conf_pct}")
            lines.append(f"- **Section:** {req.source_section} (p.{req.page_ref})")
            
            # Add SAP modules if available
            if hasattr(req, 'sap_modules') and req.sap_modules:
                sap_modules_str = ", ".join(req.sap_modules)
                lines.append(f"- **SAP Modules:** {sap_modules_str}")
            
            lines.append(f"\n{req.description}\n")
            if req.clarification_question:
                lines.append(f"> **Clarification needed:** {req.clarification_question}\n")
            if req.related_ids:
                lines.append(f"_Related: {', '.join(req.related_ids)}_\n")

    # Clarifications index
    ambig_reqs = [r for r in requirements if r.ambiguity_flag and r.clarification_question]
    if ambig_reqs:
        lines.append("\n---\n## Clarification Questions for RFP Issuer\n")
        for i, req in enumerate(ambig_reqs, start=1):
            lines.append(f"{i}. **[{req.id}]** {req.clarification_question}")

    content = "\n".join(lines)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(content)

    log.info("markdown_exported", path=out_path)
    return str(Path(out_path).resolve())


# ─────────────────────────────────────────────────────────────────────────────
# JSON export
# ─────────────────────────────────────────────────────────────────────────────

def export_json(
    requirements: list[Requirement],
    out_path: str,
    rfp_title: str = "RFP Analysis",
    include_metadata: bool = True,
    indent: int = 2,
) -> str:
    """
    Export requirements to JSON format.
    
    Args:
        requirements: List of Requirement objects to export
        out_path: Output file path for JSON
        rfp_title: Title for the analysis
        include_metadata: Whether to include metadata and summary statistics
        indent: JSON indentation level (None for compact output)
    
    Returns:
        Absolute path to the written file
    
    Example:
        >>> from outputs.exporter import export_json
        >>> export_json(requirements, "rfp_analysis.json", rfp_title="Project Alpha")
    """
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    
    # Build the JSON structure
    output = {}
    
    if include_metadata:
        # Add metadata section
        output["metadata"] = {
            "title": rfp_title,
            "generated_at": datetime.now().isoformat(),
            "total_requirements": len(requirements),
            "version": "2.0.0"
        }
        
        # Add summary statistics
        summary = {
            "by_category": {},
            "by_priority": {},
            "ambiguous_count": sum(1 for r in requirements if r.ambiguity_flag),
        }
        
        # Category breakdown
        for cat in Category:
            cat_reqs = [r for r in requirements if r.category == cat]
            if cat_reqs:
                summary["by_category"][cat.value] = {
                    "total": len(cat_reqs),
                    "must": sum(1 for r in cat_reqs if r.priority == Priority.MUST),
                    "should": sum(1 for r in cat_reqs if r.priority == Priority.SHOULD),
                    "could": sum(1 for r in cat_reqs if r.priority == Priority.COULD),
                    "wont": sum(1 for r in cat_reqs if r.priority == Priority.WONT),
                    "ambiguous": sum(1 for r in cat_reqs if r.ambiguity_flag),
                }
        
        # Priority breakdown
        for priority in Priority:
            priority_reqs = [r for r in requirements if r.priority == priority]
            if priority_reqs:
                summary["by_priority"][priority.value] = len(priority_reqs)
        
        output["summary"] = summary
    
    # Add requirements array - use Pydantic's model_dump for proper serialization
    output["requirements"] = [req.model_dump(mode="json") for req in requirements]
    
    # Add clarifications section if there are ambiguous items
    ambiguous = [r for r in requirements if r.ambiguity_flag and r.clarification_question]
    if ambiguous:
        output["clarifications_needed"] = [
            {
                "id": req.id,
                "section": req.source_section,
                "title": req.title,
                "question": req.clarification_question,
                "category": req.category.value,
            }
            for req in ambiguous
        ]
    
    # Write to file
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(output, f, indent=indent, ensure_ascii=False)
    
    log.info("json_exported", path=out_path, count=len(requirements))
    return str(Path(out_path).resolve())
