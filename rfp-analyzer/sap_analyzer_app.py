"""
SAP Mapping Analyzer - Standalone Web Application

This is a separate web application that takes RFP Analyzer output
and maps requirements to SAP modules. It runs independently and can
process JSON/Markdown/Excel output from the main RFP analyzer.

Usage:
    python sap_analyzer_app.py

Then open: http://localhost:8002
"""

import json
import os
from pathlib import Path
from typing import Optional
from datetime import datetime
import structlog
from fastapi import FastAPI, File, UploadFile, Form, HTTPException
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import uvicorn
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from agents.sap_mapping_agent import create_sap_mapping_agent

# Setup logging
log = structlog.get_logger(__name__)

app = FastAPI(
    title="SAP Requirements Mapping Analyzer",
    description="Standalone tool for mapping RFP requirements to SAP modules",
    version="1.0.0"
)

# Enable CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Store analysis results temporarily
analysis_cache = {}


class AnalysisRequest(BaseModel):
    """Request model for text-based analysis."""
    requirements: list[str]
    format: str = "json"  # json, markdown, or summary


class RFPAnalyzerOutput(BaseModel):
    """Model for RFP Analyzer JSON output."""
    requirements: list[dict]


@app.get("/", response_class=HTMLResponse)
async def root():
    """Serve the SAP Mapping Analyzer UI."""
    html_path = Path(__file__).parent / "sap_analyzer_ui.html"
    if html_path.exists():
        return HTMLResponse(content=html_path.read_text())
    
    # Inline HTML if file doesn't exist
    return HTMLResponse(content="""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>SAP Requirements Mapping Analyzer</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Arial, sans-serif;
            background: #f8f9fa;
            color: #212529;
            line-height: 1.6;
        }
        .container { max-width: 1200px; margin: 0 auto; padding: 2rem; }
        .header {
            background: linear-gradient(135deg, #185FA5 0%, #0F6E56 100%);
            color: white;
            padding: 2rem;
            border-radius: 12px;
            margin-bottom: 2rem;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        }
        .header h1 { font-size: 2rem; margin-bottom: 0.5rem; }
        .card {
            background: white;
            border-radius: 12px;
            padding: 2rem;
            box-shadow: 0 2px 4px rgba(0,0,0,0.05);
            margin-bottom: 1.5rem;
        }
        textarea {
            width: 100%;
            min-height: 200px;
            padding: 1rem;
            border: 1px solid #dee2e6;
            border-radius: 8px;
            font-family: monospace;
            font-size: 0.9rem;
            resize: vertical;
        }
        .btn {
            padding: 0.75rem 1.5rem;
            border: none;
            border-radius: 8px;
            font-size: 1rem;
            font-weight: 500;
            cursor: pointer;
            background: #185FA5;
            color: white;
            margin-top: 1rem;
        }
        .btn:hover { background: #0d4a84; }
        #results { margin-top: 2rem; }
        .stat-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 1rem;
            margin-bottom: 2rem;
        }
        .stat-card {
            background: white;
            padding: 1.5rem;
            border-radius: 8px;
            border-left: 4px solid #185FA5;
            box-shadow: 0 2px 4px rgba(0,0,0,0.05);
        }
        .stat-label { font-size: 0.85rem; color: #6c757d; text-transform: uppercase; }
        .stat-value { font-size: 2rem; font-weight: 700; color: #185FA5; margin-top: 0.25rem; }
        .module-list {
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(250px, 1fr));
            gap: 1rem;
        }
        .module-card {
            background: white;
            padding: 1rem;
            border-radius: 8px;
            border: 1px solid #dee2e6;
        }
        .module-code { font-weight: 700; color: #185FA5; font-size: 1.1rem; }
        pre {
            background: #f8f9fa;
            padding: 1rem;
            border-radius: 8px;
            overflow-x: auto;
            font-size: 0.85rem;
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>🗺️ SAP Requirements Mapping Analyzer</h1>
            <p>Analyze RFP requirements and map them to SAP modules</p>
        </div>
        
        <div class="card">
            <h2>Paste Requirements</h2>
            <p style="color: #6c757d; margin-bottom: 1rem;">
                Paste requirements (one per line) or RFP Analyzer JSON output
            </p>
            <textarea id="requirementsText" placeholder="Example:
The system shall manage accounts payable including invoice processing
The system shall support general ledger accounting with real-time postings
The system shall handle purchase order creation and goods receipt

OR paste RFP Analyzer JSON:
{&quot;requirements&quot;: [{&quot;id&quot;: &quot;FR-001&quot;, &quot;description&quot;: &quot;...&quot;}]}"></textarea>
            <button class="btn" onclick="analyzeRequirements()">Analyze Requirements</button>
            <button class="btn" onclick="exportToExcel()" style="background: #0F6E56; margin-left: 1rem;">📊 Export to Excel</button>
        </div>
        
        <div id="results" style="display: none;">
            <div class="card">
                <h2>SAP Module Mapping Results</h2>
                <div id="resultsContent"></div>
            </div>
        </div>
    </div>
    
    <script>
        async function analyzeRequirements() {
            const text = document.getElementById('requirementsText').value.trim();
            if (!text) {
                alert('Please enter requirements');
                return;
            }
            
            let requirements = [];
            
            // Try to parse as JSON first
            try {
                const json = JSON.parse(text);
                if (json.requirements && Array.isArray(json.requirements)) {
                    requirements = json.requirements.map(r => r.description || r.text || r.title || '');
                } else if (Array.isArray(json)) {
                    requirements = json.map(r => typeof r === 'string' ? r : (r.description || r.text || ''));
                }
            } catch (e) {
                // Not JSON, treat as plain text
                requirements = text.split('\\n').filter(line => line.trim());
            }
            
            if (requirements.length === 0) {
                alert('No valid requirements found');
                return;
            }
            
            // Call API
            try {
                const response = await fetch('/api/analyze', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ requirements })
                });
                
                const result = await response.json();
                displayResults(result);
            } catch (error) {
                alert('Error analyzing requirements: ' + error.message);
            }
        }
        
        function displayResults(result) {
            const resultsDiv = document.getElementById('results');
            const contentDiv = document.getElementById('resultsContent');
            
            const coverage = result.coverage_analysis;
            const modules = result.mapped_modules;
            const recommendations = result.recommendations;
            const gaps = result.gaps;
            
            let html = `
                <div class="stat-grid">
                    <div class="stat-card">
                        <div class="stat-label">Total Requirements</div>
                        <div class="stat-value">${coverage.total_requirements}</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-label">Mapped to SAP</div>
                        <div class="stat-value">${coverage.mapped_requirements}</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-label">Coverage</div>
                        <div class="stat-value">${coverage.coverage_percentage}%</div>
                    </div>
                    <div class="stat-card">
                        <div class="stat-label">SAP Modules</div>
                        <div class="stat-value">${Object.keys(modules).length}</div>
                    </div>
                </div>
                
                <h3>Identified SAP Modules</h3>
                <div class="module-list">
            `;
            
            for (const [module, reqs] of Object.entries(modules)) {
                html += `
                    <div class="module-card">
                        <div class="module-code">${module}</div>
                        <div>${reqs.length} requirement(s)</div>
                    </div>
                `;
            }
            
            html += '</div>';
            
            if (recommendations.length > 0) {
                html += '<h3>Recommendations</h3><ul>';
                recommendations.forEach(rec => {
                    html += `<li>${rec}</li>`;
                });
                html += '</ul>';
            }
            
            if (gaps.unmapped_requirements.length > 0) {
                html += '<h3>Gap Analysis</h3>';
                html += '<p>The following requirements may need custom development:</p><ul>';
                gaps.unmapped_requirements.forEach(req => {
                    html += `<li>${req}</li>`;
                });
                html += '</ul>';
            }
            
            html += '<h3>Full JSON Output</h3>';
            html += `<pre>${JSON.stringify(result, null, 2)}</pre>`;
            
            contentDiv.innerHTML = html;
            resultsDiv.style.display = 'block';
        }
        
        async function exportToExcel() {
            const text = document.getElementById('requirementsText').value.trim();
            if (!text) {
                alert('Please enter requirements first');
                return;
            }
            
            let requirements = [];
            
            // Try to parse as JSON first
            try {
                const json = JSON.parse(text);
                if (json.requirements && Array.isArray(json.requirements)) {
                    requirements = json.requirements.map(r => r.description || r.text || r.title || '');
                } else if (Array.isArray(json)) {
                    requirements = json.map(r => typeof r === 'string' ? r : (r.description || r.text || ''));
                }
            } catch (e) {
                // Not JSON, treat as plain text
                requirements = text.split('\\n').filter(line => line.trim());
            }
            
            if (requirements.length === 0) {
                alert('No valid requirements found');
                return;
            }
            
            // Call Excel export API
            try {
                const response = await fetch('/api/export-excel', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ requirements })
                });
                
                if (!response.ok) {
                    throw new Error('Export failed');
                }
                
                // Download the file
                const blob = await response.blob();
                const url = window.URL.createObjectURL(blob);
                const a = document.createElement('a');
                a.href = url;
                a.download = `SAP_Analysis_${new Date().toISOString().split('T')[0]}.xlsx`;
                document.body.appendChild(a);
                a.click();
                window.URL.revokeObjectURL(url);
                document.body.removeChild(a);
                
                alert('✅ Excel file downloaded successfully!');
            } catch (error) {
                alert('Error exporting to Excel: ' + error.message);
            }
        }
    </script>
</body>
</html>
    """)


@app.post("/api/analyze")
async def analyze_requirements(request: AnalysisRequest):
    """Analyze requirements and map to SAP modules."""
    try:
        log.info("sap_analysis_request", count=len(request.requirements))
        
        # Create SAP agent
        agent = create_sap_mapping_agent()
        
        # Analyze
        result = agent.analyze_requirements(request.requirements)
        
        log.info("sap_analysis_complete", 
                modules=len(result["mapped_modules"]),
                coverage=result["coverage_analysis"]["coverage_percentage"])
        
        return JSONResponse(content=result)
        
    except Exception as e:
        log.error("sap_analysis_error", error=str(e))
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/analyze-file")
async def analyze_file(file: UploadFile = File(...)):
    """Analyze uploaded file (JSON, Markdown, or Excel)."""
    try:
        content = await file.read()
        
        # Parse based on file type
        if file.filename.endswith('.json'):
            data = json.loads(content.decode('utf-8'))
            
            # Extract requirements from RFP Analyzer JSON
            if 'requirements' in data:
                requirements = [
                    req.get('description') or req.get('text') or req.get('title', '')
                    for req in data['requirements']
                ]
            else:
                requirements = []
                
        elif file.filename.endswith('.md'):
            # Parse markdown (simple extraction)
            text = content.decode('utf-8')
            requirements = [
                line.strip('- ').strip()
                for line in text.split('\n')
                if line.strip().startswith('-') or line.strip().startswith('*')
            ]
            
        else:
            raise HTTPException(status_code=400, detail="Unsupported file type. Use JSON or MD files.")
        
        # Analyze
        agent = create_sap_mapping_agent()
        result = agent.analyze_requirements(requirements)
        
        return JSONResponse(content=result)
        
    except Exception as e:
        log.error("file_analysis_error", error=str(e))
        raise HTTPException(status_code=500, detail=str(e))


def create_excel_report(analysis_result: dict, requirements: list[str]) -> str:
    """Create an Excel report from analysis results."""
    wb = Workbook()
    
    # Define styles
    header_fill = PatternFill(start_color="185FA5", end_color="185FA5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    subheader_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
    subheader_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Title
    ws_summary['A1'] = "SAP Requirements Mapping Analysis"
    ws_summary['A1'].font = Font(bold=True, size=16, color="185FA5")
    ws_summary.merge_cells('A1:D1')
    
    ws_summary['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws_summary.merge_cells('A2:D2')
    
    # Coverage Statistics
    row = 4
    ws_summary[f'A{row}'] = "Coverage Statistics"
    ws_summary[f'A{row}'].font = subheader_font
    ws_summary[f'A{row}'].fill = subheader_fill
    ws_summary.merge_cells(f'A{row}:B{row}')
    
    coverage = analysis_result['coverage_analysis']
    row += 1
    stats = [
        ("Total Requirements", coverage['total_requirements']),
        ("Mapped to SAP", coverage['mapped_requirements']),
        ("Coverage Percentage", f"{coverage['coverage_percentage']}%"),
        ("Identified SAP Modules", len(analysis_result['mapped_modules']))
    ]
    
    for label, value in stats:
        ws_summary[f'A{row}'] = label
        ws_summary[f'B{row}'] = value
        ws_summary[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Module Summary
    row += 2
    ws_summary[f'A{row}'] = "SAP Modules Identified"
    ws_summary[f'A{row}'].font = subheader_font
    ws_summary[f'A{row}'].fill = subheader_fill
    ws_summary.merge_cells(f'A{row}:C{row}')
    
    row += 1
    ws_summary[f'A{row}'] = "Module Code"
    ws_summary[f'B{row}'] = "Module Name"
    ws_summary[f'C{row}'] = "Requirements Count"
    for col in ['A', 'B', 'C']:
        ws_summary[f'{col}{row}'].font = header_font
        ws_summary[f'{col}{row}'].fill = header_fill
    
    module_names = {
        "FI": "Financial Accounting",
        "CO": "Controlling",
        "MM": "Materials Management",
        "SD": "Sales & Distribution",
        "PP": "Production Planning",
        "QM": "Quality Management",
        "PM": "Plant Maintenance",
        "HR/HCM": "Human Capital Management",
        "WM": "Warehouse Management",
        "PS": "Project System",
        "BW/BI": "Business Warehouse / Intelligence",
        "CRM": "Customer Relationship Management",
        "SRM": "Supplier Relationship Management",
        "SCM": "Supply Chain Management"
    }
    
    for module, reqs in analysis_result['mapped_modules'].items():
        row += 1
        ws_summary[f'A{row}'] = module
        ws_summary[f'B{row}'] = module_names.get(module, module)
        ws_summary[f'C{row}'] = len(reqs)
    
    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 40
    ws_summary.column_dimensions['C'].width = 20
    
    # Sheet 2: Module Mapping Details
    ws_mapping = wb.create_sheet("Module Mapping")
    
    ws_mapping['A1'] = "SAP Module"
    ws_mapping['B1'] = "Requirement"
    for col in ['A', 'B']:
        ws_mapping[f'{col}1'].font = header_font
        ws_mapping[f'{col}1'].fill = header_fill
    
    row = 2
    for module, reqs in sorted(analysis_result['mapped_modules'].items()):
        for req in reqs:
            ws_mapping[f'A{row}'] = module
            ws_mapping[f'B{row}'] = req
            ws_mapping[f'A{row}'].border = border
            ws_mapping[f'B{row}'].border = border
            ws_mapping[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            row += 1
    
    ws_mapping.column_dimensions['A'].width = 15
    ws_mapping.column_dimensions['B'].width = 80
    
    # Sheet 3: Recommendations
    ws_rec = wb.create_sheet("Recommendations")
    
    ws_rec['A1'] = "Implementation Recommendations"
    ws_rec['A1'].font = Font(bold=True, size=14)
    ws_rec.merge_cells('A1:B1')
    
    row = 3
    for i, rec in enumerate(analysis_result['recommendations'], 1):
        ws_rec[f'A{row}'] = f"{i}."
        ws_rec[f'B{row}'] = rec
        ws_rec[f'B{row}'].alignment = Alignment(wrap_text=True)
        row += 1
    
    ws_rec.column_dimensions['A'].width = 5
    ws_rec.column_dimensions['B'].width = 90
    
    # Sheet 4: Gap Analysis
    ws_gaps = wb.create_sheet("Gap Analysis")
    
    ws_gaps['A1'] = "Unmapped Requirements"
    ws_gaps['A1'].font = Font(bold=True, size=14)
    ws_gaps.merge_cells('A1:B1')
    
    gaps = analysis_result['gaps']
    if gaps['unmapped_requirements']:
        ws_gaps['A3'] = "The following requirements were not mapped to standard SAP modules:"
        ws_gaps.merge_cells('A3:B3')
        
        row = 5
        ws_gaps[f'A{row}'] = "#"
        ws_gaps[f'B{row}'] = "Requirement"
        for col in ['A', 'B']:
            ws_gaps[f'{col}{row}'].font = header_font
            ws_gaps[f'{col}{row}'].fill = header_fill
        
        row += 1
        for i, req in enumerate(gaps['unmapped_requirements'], 1):
            ws_gaps[f'A{row}'] = i
            ws_gaps[f'B{row}'] = req
            ws_gaps[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            row += 1
        
        row += 2
        ws_gaps[f'A{row}'] = "Note:"
        ws_gaps[f'A{row}'].font = Font(bold=True)
        row += 1
        ws_gaps[f'A{row}'] = "These requirements may need:"
        ws_gaps.merge_cells(f'A{row}:B{row}')
        row += 1
        ws_gaps[f'A{row}'] = "• Custom development"
        ws_gaps.merge_cells(f'A{row}:B{row}')
        row += 1
        ws_gaps[f'A{row}'] = "• Additional SAP modules"
        ws_gaps.merge_cells(f'A{row}:B{row}')
        row += 1
        ws_gaps[f'A{row}'] = "• Third-party integrations"
        ws_gaps.merge_cells(f'A{row}:B{row}')
    else:
        ws_gaps['A3'] = "✓ All requirements have been mapped to SAP modules!"
        ws_gaps['A3'].font = Font(bold=True, size=12, color="0F6E56")
        ws_gaps.merge_cells('A3:B3')
    
    ws_gaps.column_dimensions['A'].width = 10
    ws_gaps.column_dimensions['B'].width = 90
    
    # Save to file
    output_dir = Path("outputs")
    output_dir.mkdir(exist_ok=True)
    
    filename = f"SAP_Analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = output_dir / filename
    wb.save(filepath)
    
    return str(filepath)


@app.post("/api/export-excel")
async def export_excel(request: AnalysisRequest):
    """Export analysis results to Excel."""
    try:
        log.info("excel_export_request", count=len(request.requirements))
        
        # Create SAP agent and analyze
        agent = create_sap_mapping_agent()
        result = agent.analyze_requirements(request.requirements)
        
        # Create Excel file
        filepath = create_excel_report(result, request.requirements)
        
        log.info("excel_export_complete", filepath=filepath)
        
        # Return the file
        return FileResponse(
            filepath,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=Path(filepath).name
        )
        
    except Exception as e:
        log.error("excel_export_error", error=str(e))
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/health")
async def health_check():
    """Health check endpoint."""
    return {"status": "healthy", "service": "SAP Mapping Analyzer"}


if __name__ == "__main__":
    print("=" * 70)
    print("SAP Requirements Mapping Analyzer")
    print("=" * 70)
    print("\nStarting server on http://localhost:8002")
    print("\nThis is a standalone tool for SAP-specific RFP analysis.")
    print("It processes output from the main RFP Analyzer.\n")
    
    uvicorn.run(app, host="0.0.0.0", port=8002, log_level="info")